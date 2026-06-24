"""
Extract data from Monthly Review Meeting.xlsx using openpyxl (pure Python).
Works on any OS — no Excel installation needed.
"""
import json
import os
import sys
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

ROOT = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(ROOT, 'Ops files', 'Monthly Review Meeting.xlsx')
OUTPUT_PATH = os.path.join(ROOT, '_data_block.js')

EXCEL_EPOCH = datetime(1899, 12, 30)


def to_excel_serial(dt):
    if isinstance(dt, datetime):
        delta = dt - EXCEL_EPOCH
        return str(int(delta.days + delta.seconds / 86400))
    if isinstance(dt, date):
        delta = datetime(dt.year, dt.month, dt.day) - EXCEL_EPOCH
        return str(delta.days)
    return str(dt)


def cell_to_str(raw):
    if raw is None:
        return ''
    if isinstance(raw, (datetime, date)):
        return to_excel_serial(raw)
    return str(raw)


def extract_sheet(wb, sheet_name, max_rows, header_row=1, start_col=1, end_col=None):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    if end_col is None:
        end_col = ws.max_column or 1
    actual_max_row = min(ws.max_row or 1, max_rows + header_row)

    all_rows = list(ws.iter_rows(
        min_row=header_row, max_row=actual_max_row,
        min_col=start_col, max_col=end_col
    ))
    if not all_rows:
        return []

    headers = []
    for ci, cell in enumerate(all_rows[0]):
        h = str(cell.value).strip() if cell.value is not None else ''
        if not h:
            h = f'Col{start_col + ci}'
        headers.append(h)

    rows = []
    for row_cells in all_rows[1:]:
        obj = {}
        has_data = False
        for ci, cell in enumerate(row_cells):
            v = cell_to_str(cell.value)
            obj[headers[ci]] = v
            if v != '':
                has_data = True
        if has_data:
            rows.append(obj)
    return rows


SHEET_MAP = {
    'stockFlow':          ('19. Stock Flow', 200, 1, 1, None),
    'costingDet':         ('14. Costing Details', 1200, 1, 1, 11),
    'costingCases':       ('14. Costing Details', 1200, 1, 13, 25),
    'waitingCharges':     ('8. Waiting Charges', 1100, 1, 16, 32),
    'weeklyOrders':       ('5. Week-wise Orders Details', 200, 1, 1, None),
    'damages':            ('4. % of Damages', 200, 3, 1, None),
    'qualityIssues':      ('3. Quality Issues', 1000, 1, 1, None),
    'manualOrders':       ('6. Manual orders', 1100, 1, 1, None),
    'palletAging':        ('13. Mov. Ageing', 1200, 1, 1, None),
    'copackingWeekly':    ('29. COPACKING ORDERS-WEEK WISE', 200, 1, 1, None),
    'whDamages':          ('12. WH Handling Damages', 200, 1, 1, None),
    'inboundFlow':        ('16. Inbound Flow-25', 300, 1, 1, 4),
    'invoiceSummary':     ('24. WH Invoice Summary', 300, 1, 1, None),
    'tempReport':         ('20. Temp report', 1100, 1, 1, None),
    'subjects':           ('1. Subjects', 300, 1, 1, None),
    'hubInbound':         ('15. Hubwise Inbound Summary', 1100, 1, 1, None),
    'hubOutbound':        ('18. Hubwise Outbound Summary', 700, 1, 1, None),
    'storage':            ('10. Storage', 1100, 1, 16, 45),
    'storageHub':         ('10. Storage', 50, 1, 1, 14),
    'copackingOrders':    ('28. Co-Packing Orders', 700, 1, 1, 4),
    'coPacking':          ('26. Co-Packing', 1400, 1, 1, None),
    'customerComplaints': ('27. Customers Complaints', 1100, 1, 1, None),
    'inboundIssues':      ('25. Inbound Issues', 300, 1, 1, None),
    'inbGrnMonthly':      ('2. Inbound Plan vs GRN', 1100, 1, 8, 18),
    'expiredStock':       ('11. Expiry & Near Expire', 1400, 1, 1, None),
    'freights':           ('7. Freights', 1100, 1, 1, None),
    'outboundCases':      ('22. Outbound Summary', 1100, 1, 1, None),
    'truckTurnover':      ('9. Truck Turnover Time', 3600, 1, 1, None),
    'inboundSummary':     ('17. Inbound Summary', 300, 1, 1, None),
    'inboundDetail':      ('21. Inbound Detail', 2000, 1, 1, None),
    'ytdTrends':          ('23. YTD Trends', 400, 1, 1, None),
    'fzeSkuGw':           ("30. FZE SKU's GW Details", 100, 1, 1, None),
}


def run_extraction(excel_path=None, output_path=None):
    excel_path = excel_path or EXCEL_PATH
    output_path = output_path or OUTPUT_PATH

    if not os.path.exists(excel_path):
        return False, f'Excel file not found: {excel_path}'

    print(f'Opening {os.path.basename(excel_path)}...')
    wb = load_workbook(excel_path, read_only=True, data_only=True)

    all_data = {}
    total = len(SHEET_MAP)
    for i, (key, params) in enumerate(SHEET_MAP.items(), 1):
        sheet_name, max_rows, header_row, start_col, end_col = params
        print(f'  [{i}/{total}] Extracting: {key} ({sheet_name})...')
        rows = extract_sheet(wb, sheet_name, max_rows, header_row, start_col, end_col)
        all_data[key] = rows

    wb.close()

    result = 'var EMBEDDED_DATA=' + json.dumps(all_data, ensure_ascii=False, separators=(',', ':')) + ';'

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(result)

    size_kb = len(result) // 1024
    print(f'\nSUCCESS: Data extracted to {os.path.basename(output_path)} ({size_kb} KB)')
    return True, f'Extracted {total} datasets ({size_kb} KB)'


if __name__ == '__main__':
    ok, msg = run_extraction()
    print(msg)
    sys.exit(0 if ok else 1)
